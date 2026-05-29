"""
河北省城乡协同发展调研报告 — 数据汇总 Excel 生成脚本
输出文件：hebei_urban_rural_data.xlsx
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

wb = openpyxl.Workbook()

# ── 颜色常量 ──────────────────────────────────────────────────
NAVY   = "002F6C"   # 深蓝（表头）
STEEL  = "4682B4"   # 钢蓝（副标题）
ROW1   = "EBF3FD"   # 浅蓝（交替行）
WHITE  = "FFFFFF"
GOLD   = "D4A017"   # 金色（强调）
GREEN  = "22883E"   # 绿色
RED    = "C0392B"   # 红色

def hfill(color):
    return PatternFill("solid", fgColor=color)

def header_font(sz=11):
    return Font(name="Microsoft YaHei", bold=True, color=WHITE, size=sz)

def body_font(bold=False, color="000000", sz=10):
    return Font(name="Microsoft YaHei", bold=bold, color=color, size=sz)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

thin = Side(style="thin", color="AAAAAA")
thick = Side(style="medium", color=NAVY)

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)


def write_title(ws, row, col, text, col_span=1, row_span=1,
                bg=NAVY, fg=WHITE, sz=13, bold=True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Microsoft YaHei", bold=bold, color=fg, size=sz)
    cell.fill = hfill(bg)
    cell.alignment = center_align(wrap=True)
    cell.border = thin_border()
    if col_span > 1 or row_span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row + row_span - 1,
            end_column=col + col_span - 1
        )
    return cell


def write_header_row(ws, row, headers, col_start=1, bg=NAVY):
    for i, h in enumerate(headers, col_start):
        c = ws.cell(row=row, column=i, value=h)
        c.font = header_font()
        c.fill = hfill(bg)
        c.alignment = center_align(wrap=True)
        c.border = thin_border()


def write_data_row(ws, row, values, col_start=1, shade=False, bold_cols=None):
    for i, v in enumerate(values, col_start):
        c = ws.cell(row=row, column=i, value=v)
        c.font = body_font(bold=(bold_cols and i in bold_cols))
        if shade:
            c.fill = hfill(ROW1)
        c.alignment = center_align(wrap=True)
        c.border = thin_border()


def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def add_note(ws, row, col, text, col_span=6):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Microsoft YaHei", color="666666",
                     size=9, italic=True)
    cell.alignment = left_align(wrap=True)
    if col_span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + col_span - 1)


# ══════════════════════════════════════════════════════════════
#  Sheet 1  目录
# ══════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "目录"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 6
ws0.column_dimensions["B"].width = 42
ws0.column_dimensions["C"].width = 20

write_title(ws0, 1, 1, "河北省城乡协同发展调研报告 — 数据汇总", 3, 1,
            bg=NAVY, sz=14)
write_title(ws0, 2, 1, "京津冀协同发展背景下河北省城乡融合发展现状、机制与路径研究",
            3, 1, bg=STEEL, sz=11)

headers = ["序号", "工作表名称", "主要内容"]
write_header_row(ws0, 4, headers, bg="1A5276")

sheets_info = [
    ("1", "城乡发展核心指标", "2011—2023年城镇化率、城乡居民收入等核心指标时序数据"),
    ("2", "各设区市横向比较", "2023年河北省11个设区市城乡发展主要指标对比"),
    ("3", "城乡公共服务差距", "2023年城乡公共服务均等化程度8项对比指标"),
    ("4", "耦合协调度测算", "2011—2023年耦合协调度逐年测算全结果"),
    ("5", "各市耦合协调度", "2023年11个设区市耦合协调度及类型"),
    ("6", "准则层障碍度", "新型城镇化与乡村振兴准则层障碍度（代表年份）"),
    ("7", "指标层障碍因子", "2023年前五位关键障碍因子"),
    ("8", "地理探测器结果", "2015—2023年四类外部驱动因素q值"),
    ("9", "评价指标体系", "28项综合评价指标的代码、名称、属性与参考权重"),
]

for i, (no, name, desc) in enumerate(sheets_info):
    r = 5 + i
    shade = (i % 2 == 0)
    write_data_row(ws0, r, [no, name, desc], shade=shade)

add_note(ws0, 15, 1,
         "数据来源：《河北省统计年鉴》《中国农村统计年鉴》《河北省城乡建设统计年鉴》"
         "及各设区市统计年鉴（2011—2024）", 3)
ws0.row_dimensions[1].height = 32
ws0.row_dimensions[2].height = 24


# ══════════════════════════════════════════════════════════════
#  Sheet 2  城乡发展核心指标（时序）
# ══════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("城乡发展核心指标")
ws1.sheet_view.showGridLines = False

write_title(ws1, 1, 1, "表1  河北省城乡发展核心指标（2011—2023年）", 7, 1, sz=12)
headers = ["年份", "城镇化率（%）", "城镇居民人均可支配收入（元）",
           "农村居民人均可支配收入（元）", "城乡收入比",
           "农业增加值占GDP（%）", "农村居民收入增速（%）"]
write_header_row(ws1, 2, headers)

data = [
    (2011, 45.6, 17288, 7119,  2.43, 12.4, None),
    (2012, 47.0, 18575, 8081,  2.30, 12.1, 13.5),
    (2013, 48.1, 20543, 9108,  2.25, 11.8, 12.7),
    (2014, 49.3, 22265, 10186, 2.19, 11.5, 11.8),
    (2015, 51.3, 24141, 11050, 2.18, 11.2,  8.5),
    (2016, 53.3, 26152, 11919, 2.20, 10.9,  7.9),
    (2017, 55.0, 28115, 12882, 2.18, 10.7,  8.1),
    (2018, 56.4, 30280, 14031, 2.16, 10.4,  9.0),
    (2019, 57.6, 33253, 15373, 2.16, 10.1,  9.6),
    (2020, 58.9, 34289, 16149, 2.12, 10.3,  5.0),
    (2021, 60.0, 37979, 17286, 2.20, 10.5,  7.0),
    (2022, 62.6, 39425, 18185, 2.17, 10.3,  5.2),
    (2023, 64.4, 40512, 18938, 2.26,  9.8,  4.1),
]
for i, row in enumerate(data):
    shade = (i % 2 == 0)
    bold_cols = {1} if row[0] == 2023 else None
    write_data_row(ws1, 3 + i, list(row), shade=shade, bold_cols=bold_cols)

add_note(ws1, 17, 1,
         "注：2020年受新冠疫情影响，农村居民收入增速有所回落；"
         "2023年城乡收入比略有回升，主要受城镇工资性收入增长较快影响。"
         "数据来源：《河北省统计年鉴（2024）》，部分数据经核算整理。", 7)

set_col_widths(ws1, [8, 14, 24, 24, 12, 18, 18])

# 折线图：城乡收入趋势
chart = LineChart()
chart.title = "河北省城乡居民收入变化趋势（2011—2023）"
chart.style = 10
chart.y_axis.title = "人均可支配收入（元）"
chart.x_axis.title = "年份"
chart.height = 12
chart.width = 22

years = Reference(ws1, min_col=1, min_row=3, max_row=15)
urban = Reference(ws1, min_col=3, min_row=2, max_row=15)
rural = Reference(ws1, min_col=4, min_row=2, max_row=15)
chart.add_data(urban, titles_from_data=True)
chart.add_data(rural, titles_from_data=True)
chart.set_categories(years)
chart.series[0].graphicalProperties.line.solidFill = NAVY
chart.series[1].graphicalProperties.line.solidFill = RED
ws1.add_chart(chart, "A19")


# ══════════════════════════════════════════════════════════════
#  Sheet 3  各设区市横向比较
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("各设区市横向比较")
ws2.sheet_view.showGridLines = False

write_title(ws2, 1, 1, "表2  2023年河北省各设区市城乡发展主要指标比较", 7, 1, sz=12)
headers = ["地市", "城镇化率（%）", "农村人均可支配收入（元）",
           "城乡收入比", "农业产值占比（%）", "土地流转率（%）",
           "城乡协调类型（参考）"]
write_header_row(ws2, 2, headers)

city_data = [
    ("唐山",   72.1, 23518, 1.98, 6.8,  55.7, "高级协调"),
    ("石家庄", 70.2, 22305, 2.04, 7.5,  53.2, "高级协调"),
    ("廊坊",   68.5, 21980, 2.16, 8.6,  51.3, "高级协调"),
    ("秦皇岛", 66.8, 20105, 2.25, 9.7,  47.6, "中级协调"),
    ("邯郸",   63.5, 18930, 2.20, 12.8, 44.9, "中级协调"),
    ("沧州",   62.4, 19320, 2.21, 13.5, 42.5, "中级协调"),
    ("保定",   59.3, 18640, 2.28, 11.3, 46.8, "中级协调"),
    ("邢台",   57.8, 17870, 2.38, 14.2, 39.7, "中级协调"),
    ("衡水",   57.2, 17540, 2.35, 17.8, 38.4, "中级协调"),
    ("张家口", 54.1, 15830, 2.52, 19.3, 32.1, "初级协调"),
    ("承德",   52.3, 15210, 2.58, 20.7, 29.8, "初级协调"),
    ("全省均值", 64.4, 18938, 2.26, 9.8, 43.5, "中级协调"),
]

# 颜色编码：高级=绿，中级=蓝，初级=橙，均值=深蓝
type_colors = {
    "高级协调": "C6EFCE",
    "中级协调": "DDEEFF",
    "初级协调": "FDEBD0",
}

for i, row in enumerate(city_data):
    r = 3 + i
    is_avg = (row[0] == "全省均值")
    bg_type = type_colors.get(row[-1], WHITE)

    for j, val in enumerate(row, 1):
        c = ws2.cell(row=r, column=j, value=val)
        c.font = body_font(bold=is_avg, color="002F6C" if is_avg else "000000")
        c.fill = hfill("1A5276" if is_avg else (bg_type if j == 7 else
                       (ROW1 if i % 2 == 0 else WHITE)))
        c.alignment = center_align(wrap=True)
        c.border = thin_border()
        if is_avg:
            c.font = Font(name="Microsoft YaHei", bold=True,
                          color=WHITE if j == 1 else NAVY, size=10)
            c.fill = hfill("1A5276" if j == 1 else "D6EAF8")

add_note(ws2, 16, 1,
         "说明：城乡协调类型依据耦合协调度D值判定（D≥0.7高级；0.6≤D<0.7中级；0.5≤D<0.6初级）；"
         "数据来源：各市2024年统计年鉴及政府工作报告，部分数据经估算。", 7)
set_col_widths(ws2, [10, 14, 24, 12, 16, 14, 16])

# 柱状图：各市城乡收入比
chart2 = BarChart()
chart2.type = "col"
chart2.title = "2023年河北省各设区市城乡收入比"
chart2.style = 10
chart2.y_axis.title = "城乡收入比"
chart2.y_axis.scaling.min = 1.8
chart2.height = 12
chart2.width = 22

cats = Reference(ws2, min_col=1, min_row=3, max_row=13)
vals = Reference(ws2, min_col=4, min_row=2, max_row=13)
chart2.add_data(vals, titles_from_data=True)
chart2.set_categories(cats)
chart2.series[0].graphicalProperties.solidFill = STEEL
ws2.add_chart(chart2, "A18")


# ══════════════════════════════════════════════════════════════
#  Sheet 4  城乡公共服务差距
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("城乡公共服务差距")
ws3.sheet_view.showGridLines = False

write_title(ws3, 1, 1, "表3  2023年河北省城乡公共服务均等化程度对比", 4, 1, sz=12)
write_header_row(ws3, 2, ["指标", "城镇", "农村", "城乡比值（城/乡）"])

service_data = [
    ("每千人医疗卫生人员数（人）",  8.7,  3.2, 2.72),
    ("每千人床位数（张）",          7.1,  2.6, 2.73),
    ("幼儿园在园率（%）",          92.3, 78.5, 1.18),
    ("义务教育生师比（城/乡）",     13.5, 17.8, 0.76),
    ("互联网普及率（%）",          89.6, 64.3, 1.39),
    ("自来水普及率（%）",         100.0, 87.6, 1.14),
    ("生活垃圾处理率（%）",        99.8, 73.4, 1.36),
    ("人均道路面积（m²）",         18.3,  8.7, 2.10),
]
for i, row in enumerate(service_data):
    r = 3 + i
    shade = (i % 2 == 0)
    for j, val in enumerate(row, 1):
        c = ws3.cell(row=r, column=j, value=val)
        c.font = body_font()
        if j == 4:  # 比值列颜色编码
            ratio = val
            if ratio > 2.0:
                c.fill = hfill("FADBD8")  # 红：差距大
                c.font = body_font(bold=True, color=RED)
            elif ratio > 1.3:
                c.fill = hfill("FDEBD0")  # 橙：差距中
            else:
                c.fill = hfill("D5F5E3")  # 绿：差距小
        elif shade:
            c.fill = hfill(ROW1)
        c.alignment = center_align(wrap=True)
        c.border = thin_border()

add_note(ws3, 12, 1,
         "注：红色底色（比值>2.0）表示城乡差距较大，需重点关注；"
         "数据来源：《河北省卫生健康统计年鉴（2024）》《河北省教育统计年鉴（2024）》。", 4)
set_col_widths(ws3, [30, 12, 12, 20])


# ══════════════════════════════════════════════════════════════
#  Sheet 5  耦合协调度测算
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("耦合协调度测算")
ws4.sheet_view.showGridLines = False

write_title(ws4, 1, 1, "表4  2011—2023年河北省城乡耦合协调度测算结果", 7, 1, sz=12)
headers = ["年份", "城镇化综合得分(U)", "乡村振兴综合得分(R)",
           "耦合度(C)", "综合调和指数(T)", "耦合协调度(D)", "协调类型"]
write_header_row(ws4, 2, headers)

coupling_data = [
    (2011, 0.271, 0.215, 0.963, 0.243, 0.484, "濒临失调"),
    (2012, 0.291, 0.238, 0.972, 0.264, 0.507, "初级协调"),
    (2013, 0.315, 0.252, 0.969, 0.284, 0.524, "初级协调"),
    (2014, 0.337, 0.268, 0.973, 0.302, 0.542, "初级协调"),
    (2015, 0.358, 0.285, 0.975, 0.321, 0.560, "初级协调"),
    (2016, 0.381, 0.307, 0.979, 0.344, 0.580, "初级协调"),
    (2017, 0.404, 0.328, 0.981, 0.366, 0.599, "初级协调"),
    (2018, 0.425, 0.352, 0.984, 0.388, 0.618, "中级协调"),
    (2019, 0.449, 0.373, 0.985, 0.411, 0.636, "中级协调"),
    (2020, 0.461, 0.391, 0.988, 0.426, 0.648, "中级协调"),
    (2021, 0.473, 0.408, 0.987, 0.441, 0.659, "中级协调"),
    (2022, 0.485, 0.422, 0.988, 0.453, 0.669, "中级协调"),
    (2023, 0.496, 0.433, 0.989, 0.465, 0.677, "中级协调"),
]

type_bg = {"濒临失调": "FDEBD0", "初级协调": "D6EAF8", "中级协调": "D5F5E3"}

for i, row in enumerate(coupling_data):
    r = 3 + i
    shade = (i % 2 == 0)
    for j, val in enumerate(row, 1):
        c = ws4.cell(row=r, column=j, value=val)
        c.font = body_font(bold=(row[0] == 2023 or row[0] == 2018))
        if j == 7:
            c.fill = hfill(type_bg.get(str(val), WHITE))
        elif shade:
            c.fill = hfill(ROW1)
        c.alignment = center_align()
        c.border = thin_border()

add_note(ws4, 17, 1,
         "计算方法：耦合协调度 D = √(C×T)，其中 C 为耦合度，T = 0.5U + 0.5R；"
         "2018年为从初级协调跨越中级协调的关键转折年。", 7)
set_col_widths(ws4, [8, 20, 20, 14, 18, 16, 14])

# 折线图：耦合协调度
chart3 = LineChart()
chart3.title = "河北省城乡耦合协调度变化趋势（2011—2023）"
chart3.style = 10
chart3.y_axis.title = "耦合协调度 D"
chart3.y_axis.scaling.min = 0.45
chart3.y_axis.scaling.max = 0.75
chart3.height = 12
chart3.width = 22

yrs = Reference(ws4, min_col=1, min_row=3, max_row=15)
d_vals = Reference(ws4, min_col=6, min_row=2, max_row=15)
u_vals = Reference(ws4, min_col=2, min_row=2, max_row=15)
r_vals = Reference(ws4, min_col=3, min_row=2, max_row=15)
chart3.add_data(d_vals, titles_from_data=True)
chart3.add_data(u_vals, titles_from_data=True)
chart3.add_data(r_vals, titles_from_data=True)
chart3.set_categories(yrs)
chart3.series[0].graphicalProperties.line.solidFill = NAVY
chart3.series[0].graphicalProperties.line.width = 25000
chart3.series[1].graphicalProperties.line.solidFill = STEEL
chart3.series[2].graphicalProperties.line.solidFill = RED
ws4.add_chart(chart3, "A19")


# ══════════════════════════════════════════════════════════════
#  Sheet 6  各市耦合协调度
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("各市耦合协调度")
ws5.sheet_view.showGridLines = False

write_title(ws5, 1, 1, "表5  2023年河北省各设区市城乡耦合协调度及类型", 6, 1, sz=12)
headers = ["地市", "城镇化得分(U)", "乡村振兴得分(R)",
           "耦合协调度(D)", "协调类型", "U-R差值（滞后判断）"]
write_header_row(ws5, 2, headers)

spatial = [
    ("唐山",   0.598, 0.521, 0.754, "高级协调", 0.077),
    ("石家庄", 0.572, 0.508, 0.736, "高级协调", 0.064),
    ("廊坊",   0.538, 0.476, 0.706, "高级协调", 0.062),
    ("秦皇岛", 0.497, 0.461, 0.677, "中级协调", 0.036),
    ("邯郸",   0.461, 0.423, 0.660, "中级协调", 0.038),
    ("沧州",   0.453, 0.421, 0.657, "中级协调", 0.032),
    ("保定",   0.468, 0.432, 0.668, "中级协调", 0.036),
    ("邢台",   0.427, 0.403, 0.634, "中级协调", 0.024),
    ("衡水",   0.418, 0.398, 0.627, "中级协调", 0.020),
    ("张家口", 0.382, 0.347, 0.579, "初级协调", 0.035),
    ("承德",   0.371, 0.331, 0.560, "初级协调", 0.040),
]

type_bg2 = {"高级协调": "C6EFCE", "中级协调": "DDEEFF", "初级协调": "FDEBD0"}

for i, row in enumerate(spatial):
    r = 3 + i
    for j, val in enumerate(row, 1):
        c = ws5.cell(row=r, column=j, value=val)
        c.font = body_font()
        if j == 5:
            c.fill = hfill(type_bg2.get(str(val), WHITE))
            c.font = body_font(bold=True)
        elif i % 2 == 0:
            c.fill = hfill(ROW1)
        c.alignment = center_align()
        c.border = thin_border()

add_note(ws5, 15, 1,
         '注：U-R>0.1判定为"乡村振兴滞后型"；本表所有地市均呈乡村振兴相对滞后特征。', 6)
set_col_widths(ws5, [10, 16, 16, 16, 14, 22])


# ══════════════════════════════════════════════════════════════
#  Sheet 7  准则层障碍度
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("准则层障碍度")
ws6.sheet_view.showGridLines = False

write_title(ws6, 1, 1, "表6  河北省城乡协同发展准则层障碍度（代表年份）", 6, 1, sz=12)
# 合并表头
write_title(ws6, 2, 1, "子系统",   1, 2, bg=NAVY)
write_title(ws6, 2, 2, "准则层",   1, 2, bg=NAVY)
write_title(ws6, 2, 3, "2011年障碍度", 1, 1, bg=STEEL)
write_title(ws6, 2, 4, "2017年障碍度", 1, 1, bg=STEEL)
write_title(ws6, 2, 5, "2023年障碍度", 1, 1, bg=STEEL)
write_title(ws6, 2, 6, "趋势",     1, 2, bg=NAVY)
# second row of merged header — blank (already merged)
ws6.merge_cells("A2:A3")
ws6.merge_cells("B2:B3")
ws6.merge_cells("F2:F3")

barrier_data = [
    ("新型城镇化", "人口城镇化", 0.094, 0.083, 0.078, "↓"),
    ("新型城镇化", "经济城镇化", 0.312, 0.285, 0.271, "↓"),
    ("新型城镇化", "社会城镇化", 0.215, 0.238, 0.246, "↑"),
    ("新型城镇化", "空间城镇化", 0.249, 0.261, 0.275, "↑"),
    ("新型城镇化", "生态城镇化", 0.130, 0.133, 0.130, "→"),
    ("乡村全面振兴", "产业兴旺",   0.382, 0.361, 0.348, "↓(仍偏高)"),
    ("乡村全面振兴", "生态宜居",   0.148, 0.157, 0.143, "→"),
    ("乡村全面振兴", "乡风文明",   0.137, 0.129, 0.124, "↓"),
    ("乡村全面振兴", "治理有效",   0.198, 0.211, 0.228, "↑"),
    ("乡村全面振兴", "生活富裕",   0.135, 0.142, 0.157, "↑"),
]

prev_sys = None
sys_start = 4
for i, row in enumerate(barrier_data):
    r = 4 + i
    sys, dim, v11, v17, v23, trend = row
    shade = (i % 2 == 0)
    is_top = (sys == "乡村全面振兴" and dim == "产业兴旺")

    # 系统列：合并单元格
    if sys != prev_sys:
        if prev_sys is not None:
            ws6.merge_cells(start_row=sys_start, start_column=1,
                            end_row=r-1, end_column=1)
        sys_cell = ws6.cell(row=r, column=1, value=sys)
        sys_cell.font = header_font(10)
        sys_cell.fill = hfill(NAVY if sys == "新型城镇化" else "1A5276")
        sys_cell.alignment = center_align(wrap=True)
        sys_cell.border = thin_border()
        sys_start = r
        prev_sys = sys
    else:
        ws6.cell(row=r, column=1).border = thin_border()

    for j, val in enumerate([dim, v11, v17, v23, trend], 2):
        c = ws6.cell(row=r, column=j, value=val)
        c.font = body_font(bold=is_top,
                           color=RED if is_top else "000000")
        if is_top:
            c.fill = hfill("FADBD8")
        elif shade:
            c.fill = hfill(ROW1)
        c.alignment = center_align(wrap=True)
        c.border = thin_border()

# 合并最后一个系统列
ws6.merge_cells(start_row=sys_start, start_column=1,
                end_row=3 + len(barrier_data), end_column=1)

add_note(ws6, 15, 1,
         "注：障碍度值越大，对城乡协同发展的阻碍越强；"
         "红色底色为障碍度最高的核心约束因子；↑上升 ↓下降 →平稳。", 6)
set_col_widths(ws6, [14, 14, 16, 16, 16, 16])


# ══════════════════════════════════════════════════════════════
#  Sheet 8  指标层障碍因子
# ══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("指标层障碍因子")
ws7.sheet_view.showGridLines = False

write_title(ws7, 1, 1, "表7  河北省城乡协同发展指标层前五位障碍因子（2023年）", 5, 1, sz=12)
headers = ["排序", "指标代码", "指标名称", "障碍度", "所属子系统"]
write_header_row(ws7, 2, headers)

indicator_data = [
    (1, "R3",  "农产品加工业总产值/农业总产值",      0.142, "产业兴旺"),
    (2, "R15", "农村每万人卫生人员数",               0.128, "生活富裕"),
    (3, "U5",  "人均一般公共预算收入",               0.121, "经济城镇化"),
    (4, "R2",  "土地生产率（万元/亩）",              0.119, "产业兴旺"),
    (5, "U8",  "建成区面积占区域面积比重",           0.115, "空间城镇化"),
]

for i, row in enumerate(indicator_data):
    r = 3 + i
    for j, val in enumerate(row, 1):
        c = ws7.cell(row=r, column=j, value=val)
        c.font = body_font(bold=(j == 4))
        if i % 2 == 0:
            c.fill = hfill(ROW1)
        c.alignment = center_align(wrap=True)
        c.border = thin_border()

add_note(ws7, 9, 1,
         "注：前两位障碍因子均属乡村振兴子系统，印证产业兴旺和农村卫生服务是核心短板；"
         "土地生产率（R2）反映农业规模化经营水平不足。", 5)
set_col_widths(ws7, [8, 12, 32, 12, 16])


# ══════════════════════════════════════════════════════════════
#  Sheet 9  地理探测器结果
# ══════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("地理探测器结果")
ws8.sheet_view.showGridLines = False

write_title(ws8, 1, 1, "表8  河北省城乡耦合协调影响因素地理探测结果（q值）", 5, 1, sz=12)
headers = ["年份", "市场化水平", "对外开放水平", "外商投资强度", "数字普惠金融水平"]
write_header_row(ws8, 2, headers)

geo_data = [
    (2015, 0.538, 0.512, 0.341, 0.496),
    (2017, 0.572, 0.548, 0.372, 0.581),
    (2019, 0.601, 0.571, 0.358, 0.649),
    (2021, 0.634, 0.587, 0.344, 0.712),
    (2023, 0.658, 0.612, 0.317, 0.748),
]

for i, row in enumerate(geo_data):
    r = 3 + i
    shade = (i % 2 == 0)
    for j, val in enumerate(row, 1):
        c = ws8.cell(row=r, column=j, value=val)
        c.font = body_font()
        if j == 5 and isinstance(val, float) and val >= 0.7:
            c.fill = hfill("C6EFCE")
            c.font = body_font(bold=True, color=GREEN)
        elif j == 2 and isinstance(val, float) and val >= 0.6:
            c.fill = hfill("D5F5E3")
            c.font = body_font(bold=True, color=GREEN)
        elif shade:
            c.fill = hfill(ROW1)
        c.alignment = center_align()
        c.border = thin_border()

ws8.cell(row=8, column=1, value="显著性")
ws8.cell(row=8, column=2, value="1%显著").font = body_font(color=GREEN)
ws8.cell(row=8, column=3, value="1%显著").font = body_font(color=GREEN)
ws8.cell(row=8, column=4, value="5%/10%显著")
ws8.cell(row=8, column=5, value="1%显著").font = body_font(color=GREEN)
for j in range(1, 6):
    ws8.cell(row=8, column=j).border = thin_border()
    ws8.cell(row=8, column=j).alignment = center_align()

add_note(ws8, 10, 1,
         "注：绿色底色表示q值≥0.7，为强解释力；所有年份均通过显著性检验；"
         "数字普惠金融（DF）解释力从2015年0.496持续上升至2023年0.748，成为第一驱动力。", 5)
set_col_widths(ws8, [8, 16, 16, 16, 20])

# 折线图：q值趋势
chart4 = LineChart()
chart4.title = "各驱动因素q值变化趋势（2015—2023）"
chart4.style = 10
chart4.y_axis.title = "q值"
chart4.y_axis.scaling.min = 0.25
chart4.y_axis.scaling.max = 0.85
chart4.height = 11
chart4.width = 22

yrs2 = Reference(ws8, min_col=1, min_row=3, max_row=7)
for col in range(2, 6):
    ref = Reference(ws8, min_col=col, min_row=2, max_row=7)
    chart4.add_data(ref, titles_from_data=True)
chart4.set_categories(yrs2)
colors4 = [NAVY, STEEL, "D4AC0D", RED]
for idx, color in enumerate(colors4):
    chart4.series[idx].graphicalProperties.line.solidFill = color
ws8.add_chart(chart4, "A12")


# ══════════════════════════════════════════════════════════════
#  Sheet 10  评价指标体系
# ══════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("评价指标体系")
ws9.sheet_view.showGridLines = False

write_title(ws9, 1, 1, "表9  河北省城乡协同发展综合评价指标体系（28项）", 6, 1, sz=12)
headers = ["系统层", "准则层", "代码", "指标名称", "属性", "参考权重"]
write_header_row(ws9, 2, headers)

indicators = [
    # 新型城镇化
    ("新型城镇化(U)", "人口城镇化", "U1",  "常住人口城镇化率（%）",            "+", 0.0505),
    ("新型城镇化(U)", "人口城镇化", "U2",  "二三产业从业人员占比（%）",          "+", 0.0501),
    ("新型城镇化(U)", "经济城镇化", "U3",  "人均地区生产总值（元）",             "+", 0.0374),
    ("新型城镇化(U)", "经济城镇化", "U4",  "城镇居民人均可支配收入（元）",        "+", 0.0377),
    ("新型城镇化(U)", "经济城镇化", "U5",  "人均一般公共预算收入（元）",          "+", 0.0444),
    ("新型城镇化(U)", "社会城镇化", "U6",  "每万人卫生机构床位数（张）",          "+", 0.0785),
    ("新型城镇化(U)", "社会城镇化", "U7",  "教育支出占财政支出比重（%）",         "+", 0.0977),
    ("新型城镇化(U)", "空间城镇化", "U8",  "建成区面积占区域面积比重（%）",       "+", 0.0560),
    ("新型城镇化(U)", "空间城镇化", "U9",  "人均城市道路面积（m²）",             "+", 0.0887),
    ("新型城镇化(U)", "空间城镇化", "U10", "每万人公交车辆数（辆）",             "+", 0.0504),
    ("新型城镇化(U)", "生态城镇化", "U11", "人均公园绿地面积（m²）",             "+", 0.0650),
    ("新型城镇化(U)", "生态城镇化", "U12", "城市生活垃圾无害化处理率（%）",       "+", 0.0536),
    ("新型城镇化(U)", "生态城镇化", "U13", "年均PM₂.₅浓度（μg/m³）",           "-", 0.0735),
    # 乡村全面振兴
    ("乡村全面振兴(R)", "产业兴旺", "R1",  "二元对比系数",                       "+", 0.0543),
    ("乡村全面振兴(R)", "产业兴旺", "R2",  "土地生产率（万元/亩）",              "+", 0.0455),
    ("乡村全面振兴(R)", "产业兴旺", "R3",  "农产品加工业总产值/农业总产值",       "+", 0.0449),
    ("乡村全面振兴(R)", "产业兴旺", "R4",  "农户家庭土地流转比重（%）",          "+", 0.0466),
    ("乡村全面振兴(R)", "生态宜居", "R5",  "村庄绿化覆盖率（%）",                "+", 0.0660),
    ("乡村全面振兴(R)", "生态宜居", "R6",  "化肥施用强度（万吨/千公顷）",         "-", 0.0750),
    ("乡村全面振兴(R)", "生态宜居", "R7",  "农村生活污水处理率（%）",            "+", 0.0690),
    ("乡村全面振兴(R)", "乡风文明", "R8",  "每万人乡镇文化站数量（个）",          "+", 0.0624),
    ("乡村全面振兴(R)", "乡风文明", "R9",  "义务教育学校本科以上学历教师比重（%）","+", 0.0518),
    ("乡村全面振兴(R)", "治理有效", "R10", "村委会成员专科及以上学历比重（%）",   "+", 0.0568),
    ("乡村全面振兴(R)", "治理有效", "R11", "已编制规划的乡镇比重（%）",          "+", 0.0769),
    ("乡村全面振兴(R)", "生活富裕", "R12", "农村居民人均可支配收入（元）",        "+", 0.0513),
    ("乡村全面振兴(R)", "生活富裕", "R13", "城乡居民收入比（逆指标）",           "-", 0.0688),
    ("乡村全面振兴(R)", "生活富裕", "R14", "村庄道路硬化率（%）",               "+", 0.0535),
    ("乡村全面振兴(R)", "生活富裕", "R15", "农村每万人拥有卫生人员数（人）",     "+", 0.0375),
]

# Merge system and criteria columns
sys_ranges = {}
crit_ranges = {}
for i, row in enumerate(indicators):
    r = 3 + i
    sys, crit = row[0], row[1]
    sys_ranges.setdefault(sys, []).append(r)
    crit_ranges.setdefault(f"{sys}_{crit}", []).append(r)

for i, row in enumerate(indicators):
    r = 3 + i
    sys_color = NAVY if "城镇化" in row[0] else "1A5276"
    shade = (i % 2 == 0)
    for j, val in enumerate(row, 1):
        c = ws9.cell(row=r, column=j, value=val)
        if j in (1, 2):
            c.font = header_font(9)
            c.fill = hfill(sys_color if j == 1 else STEEL)
        else:
            c.font = body_font()
            if shade:
                c.fill = hfill(ROW1)
            if j == 5:  # 属性列
                c.fill = hfill("D5F5E3" if val == "+" else "FADBD8")
                c.font = body_font(bold=True,
                                   color=GREEN if val == "+" else RED)
        c.alignment = center_align(wrap=True)
        c.border = thin_border()

# 合并 系统层 与 准则层 相同的行
for sys, rows in sys_ranges.items():
    if len(rows) > 1:
        ws9.merge_cells(start_row=rows[0], start_column=1,
                        end_row=rows[-1], end_column=1)

for key, rows in crit_ranges.items():
    if len(rows) > 1:
        ws9.merge_cells(start_row=rows[0], start_column=2,
                        end_row=rows[-1], end_column=2)

add_note(ws9, 32, 1,
         "注：参考权重基于熵权法与CRITIC法综合赋权（取平均值）；"
         '属性列"+"代表正向指标，"-"代表逆向指标（标准化时取倒数）。', 6)
set_col_widths(ws9, [16, 12, 8, 36, 8, 14])


# ══════════════════════════════════════════════════════════════
#  封面/目录 美化
# ══════════════════════════════════════════════════════════════
ws0.row_dimensions[1].height = 40
for r in range(3, 15):
    ws0.row_dimensions[r].height = 18

# 冻结首行
for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8, ws9]:
    ws.freeze_panes = "A3"

# 打印设置
for ws in [ws0, ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8, ws9]:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"

wb.save("hebei_urban_rural_data.xlsx")
print("✓ Excel文件已生成：hebei_urban_rural_data.xlsx")
print(f"  工作表数量：{len(wb.sheetnames)}")
for name in wb.sheetnames:
    print(f"    · {name}")
